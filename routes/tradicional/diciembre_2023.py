from flask import Flask, Blueprint, render_template, request, send_file
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import asyncio

novedad_dic2023 = Blueprint("novedad_dic2023", __name__)
app = Flask(__name__)  # Definir la instancia de Flask


async def comparar_archivos(archivo_A, archivo_B):
    datos_A = pd.read_excel(archivo_A, sheet_name="TAT")
    datos_B = pd.read_excel(archivo_B, sheet_name="TAT")
    datos_X = pd.read_excel(archivo_A, sheet_name="ICH")
    datos_Z = pd.read_excel(archivo_B, sheet_name="ICH")

    # Guardar las valores SO y Impactos actuales en las columnas
    datos_B["V-Cuota"] = datos_B[" CUOTA VENTA TOTAL DICIEMBRE"]
    datos_B["V-ResultCuota"] = datos_B["RESULTADO VENTA TOTAL DICIEMBRE"]
    datos_B["V-Diapers"] = datos_B[" CUOTA IMPACTOS DIAPER Y DIAPER PANTS 100"]
    datos_B["V-ResultDiapers"] = datos_B["RESULTADO  IMPACTOS DIAPER Y DIAPER PANTS 100 DICIEMBRE"]
    datos_B["V-Category"] = datos_B["CATEGORÍA"]
    
    datos_Z["V-Cuota"] = datos_Z[" CUOTA VENTA TOTAL DICIEMBRE"]
    datos_Z["V-ResultCuota"] = datos_Z["RESULTADO VENTA TOTAL DICIEMBRE"]
    datos_Z["V-Diapers"] = datos_Z[" CUOTA MARCA FOCO DIAPER Y DIAPER PANTS 100"]
    datos_Z["V-ResultDiapers"] = datos_Z["RESULTADO CUOTA DIAPER Y DIAPER PANTS 100"]
    datos_Z["V-DiapersImpact"] = datos_Z[" CUOTA IMPACTOS DIAPER Y DIAPER PANTS 100"]
    datos_Z["V-ResultDiapersImpact"] = datos_Z["RESULTADO IMPACTOS DIAPER Y DIAPER PANTS 100"]
    datos_Z["V-Category"] = datos_Z["CATEGORÍA"]

    async def procesar_datos():
        # Logica para Datos_A y Datos_B
        for idx, fila_A in datos_A.iterrows():
            id_persona = fila_A["CEDULA"]
            
            cuota_A = fila_A[" CUOTA VENTA TOTAL DICIEMBRE"]
            resultcuota_A = fila_A["RESULTADO VENTA TOTAL DICIEMBRE"]
            
            diapers_A = fila_A[" CUOTA IMPACTOS DIAPER Y DIAPER PANTS 100"]
            resultdiapers_A = fila_A["RESULTADO  IMPACTOS DIAPER Y DIAPER PANTS 100 DICIEMBRE"]
            
            category_A = fila_A["CATEGORÍA"]

            if id_persona in datos_B["CEDULA"].values:
                cuota_B = datos_B.loc[datos_B["CEDULA"] == id_persona, " CUOTA VENTA TOTAL DICIEMBRE"].values[0]
                resultcuota_B = datos_B.loc[datos_B["CEDULA"] == id_persona, "RESULTADO VENTA TOTAL DICIEMBRE"].values[0]
                
                diapers_B = datos_B.loc[datos_B["CEDULA"] == id_persona, " CUOTA IMPACTOS DIAPER Y DIAPER PANTS 100"].values[0]
                resultdiapers_B = datos_B.loc[datos_B["CEDULA"] == id_persona, "RESULTADO  IMPACTOS DIAPER Y DIAPER PANTS 100 DICIEMBRE"].values[0]

                category_B = datos_B.loc[datos_B["CEDULA"] == id_persona, "CATEGORÍA"].values[0]

                if cuota_A != cuota_B:
                    datos_B.loc[datos_B["CEDULA"] == id_persona, " CUOTA VENTA TOTAL DICIEMBRE"] = cuota_A
                if resultcuota_A != resultcuota_B:
                    datos_B.loc[datos_B["CEDULA"] == id_persona, "RESULTADO VENTA TOTAL DICIEMBRE"] = resultcuota_A
                    
                if diapers_A != diapers_B:
                    datos_B.loc[datos_B["CEDULA"] == id_persona, " CUOTA IMPACTOS DIAPER Y DIAPER PANTS 100"] = diapers_A
                if resultdiapers_A != resultdiapers_B:
                    datos_B.loc[datos_B["CEDULA"] == id_persona, "RESULTADO  IMPACTOS DIAPER Y DIAPER PANTS 100 DICIEMBRE"] = resultdiapers_A
                    
                if category_A != category_B:
                    datos_B.loc[datos_B["CEDULA"] == id_persona, "CATEGORÍA"] = category_A    
            else:
                datos_B = pd.concat([datos_B, fila_A.to_frame().T], ignore_index=True)

            datos_B["R-Cuota"] = "Igual"
            datos_B["R-ResultCuota"] = "Igual"
            datos_B["R-Diapers"] = "Igual"
            datos_B["R-ResultDiapers"] = "Igual"
            datos_B["R-Category"] = "Igual"

            datos_B.loc[datos_B["V-Cuota"] != datos_B[" CUOTA VENTA TOTAL DICIEMBRE"], "R-Cuota"] = "Hubo cambio"
            datos_B.loc[datos_B["V-ResultCuota"] != datos_B["RESULTADO VENTA TOTAL DICIEMBRE"], "R-ResultCuota"] = "Hubo cambio"

            datos_B.loc[datos_B["V-Diapers"] != datos_B[" CUOTA IMPACTOS DIAPER Y DIAPER PANTS 100"], "R-Diapers"] = "Hubo cambio"
            datos_B.loc[datos_B["V-ResultDiapers"] != datos_B["RESULTADO  IMPACTOS DIAPER Y DIAPER PANTS 100 DICIEMBRE"], "R-ResultDiapers"] = "Hubo cambio"

            datos_B.loc[datos_B["V-Category"] != datos_B["CATEGORÍA"], "R-Category"] = "Hubo cambio"

        # Logica para Datos_X y Datos_Z
        for idx, fila_X in datos_X.iterrows():
            id_persona = fila_X["CEDULA"]
            
            cuota_X = fila_X[" CUOTA VENTA TOTAL DICIEMBRE"]
            resultcuota_X = fila_X["RESULTADO VENTA TOTAL DICIEMBRE"]
            
            diapers_X = fila_X[" CUOTA MARCA FOCO DIAPER Y DIAPER PANTS 100"]
            resultdiapers_X = fila_X["RESULTADO CUOTA DIAPER Y DIAPER PANTS 100"]
            
            diapersimpact_X = fila_X[" CUOTA IMPACTOS DIAPER Y DIAPER PANTS 100"]
            resultdiapersimpact_X = fila_X["RESULTADO IMPACTOS DIAPER Y DIAPER PANTS 100"]
            
            category_X = fila_X["CATEGORÍA"]

            if id_persona in datos_Z["CEDULA"].values:
                cuota_Z = datos_Z.loc[datos_Z["CEDULA"] == id_persona, " CUOTA VENTA TOTAL DICIEMBRE"].values[0]
                resultcuota_Z = datos_Z.loc[datos_Z["CEDULA"] == id_persona, "RESULTADO VENTA TOTAL DICIEMBRE"].values[0]

                diapers_Z = datos_Z.loc[datos_Z["CEDULA"] == id_persona, " CUOTA MARCA FOCO DIAPER Y DIAPER PANTS 100"].values[0]
                resultdiapers_Z = datos_Z.loc[datos_Z["CEDULA"] == id_persona, "RESULTADO CUOTA DIAPER Y DIAPER PANTS 100"].values[0]

                diapersimpact_Z = datos_Z.loc[datos_Z["CEDULA"] == id_persona, " CUOTA IMPACTOS DIAPER Y DIAPER PANTS 100"].values[0]
                resultdiapersimpact_Z = datos_Z.loc[datos_Z["CEDULA"] == id_persona, "RESULTADO IMPACTOS DIAPER Y DIAPER PANTS 100"].values[0]

                category_Z = datos_Z.loc[datos_Z["CEDULA"] == id_persona, "CATEGORÍA"].values[0]

                if cuota_X != cuota_Z:
                    datos_Z.loc[datos_Z["CEDULA"] == id_persona, " CUOTA VENTA TOTAL DICIEMBRE"] = cuota_X
                if resultcuota_X != resultcuota_Z:
                    datos_Z.loc[datos_Z["CEDULA"] == id_persona, "RESULTADO VENTA TOTAL DICIEMBRE"] = resultcuota_X
                    
                if diapers_X != diapers_Z:
                    datos_Z.loc[datos_Z["CEDULA"] == id_persona, " CUOTA MARCA FOCO DIAPER Y DIAPER PANTS 100"] = diapers_X
                if resultdiapers_X != resultdiapers_Z:
                    datos_Z.loc[datos_Z["CEDULA"] == id_persona, "RESULTADO CUOTA DIAPER Y DIAPER PANTS 100"] = resultdiapers_X

                if diapersimpact_X != diapersimpact_Z:
                    datos_Z.loc[datos_Z["CEDULA"] == id_persona, " CUOTA IMPACTOS DIAPER Y DIAPER PANTS 100"] = diapersimpact_X
                if resultdiapersimpact_X != resultdiapersimpact_Z:
                    datos_Z.loc[datos_Z["CEDULA"] == id_persona, "RESULTADO IMPACTOS DIAPER Y DIAPER PANTS 100"] = resultdiapersimpact_X
                    
                if category_X != category_Z:
                    datos_Z.loc[datos_Z["CEDULA"] == id_persona, "CATEGORÍA"] = category_X    
            else:
                datos_Z = pd.concat([datos_Z, fila_X.to_frame().T], ignore_index=True)

        datos_Z["R-Cuota"] = "Igual"
        datos_Z["R-ResultCuota"] = "Igual"
        datos_Z["R-Diapers"] = "Igual"
        datos_Z["R-ResultDiapers"] = "Igual"
        datos_Z["R-DiapersImpact"] = "Igual"
        datos_Z["R-ResultDiapersImpact"] = "Igual"
        datos_Z["R-Category"] = "Igual"

        datos_Z.loc[datos_Z["V-Cuota"] != datos_Z[" CUOTA VENTA TOTAL DICIEMBRE"], "R-Cuota"] = "Hubo cambio"
        datos_Z.loc[datos_Z["V-ResultCuota"] != datos_Z["RESULTADO VENTA TOTAL DICIEMBRE"], "R-ResultCuota"] = "Hubo cambio"
        
        datos_Z.loc[datos_Z["V-Diapers"] != datos_Z[" CUOTA MARCA FOCO DIAPER Y DIAPER PANTS 100"], "R-Diapers"] = "Hubo cambio"
        datos_Z.loc[datos_Z["V-ResultDiapers"] != datos_Z["RESULTADO CUOTA DIAPER Y DIAPER PANTS 100"], "R-ResultDiapers"] = "Hubo cambio"

        datos_Z.loc[datos_Z["V-DiapersImpact"] != datos_Z[" CUOTA IMPACTOS DIAPER Y DIAPER PANTS 100"], "R-DiapersImpact"] = "Hubo cambio"
        datos_Z.loc[datos_Z["V-ResultDiapersImpact"] != datos_Z["RESULTADO IMPACTOS DIAPER Y DIAPER PANTS 100"], "R-ResultDiapersImpact"] = "Hubo cambio"

        datos_Z.loc[datos_Z["V-Category"] != datos_Z["CATEGORÍA"], "R-Category"] = "Hubo cambio"
        
        await procesar_datos()
    return datos_B, datos_Z

@novedad_dic2023.route("/")
def novedades():
    return render_template("/novedades/novedades.html")


@novedad_dic2023.route("/procesar_archivos", methods=["POST"])
async def procesar_archivos():
    archivo_A = request.files["archivo_A"]
    archivo_B = request.files["archivo_B"]

    datos_B, datos_Z = await comparar_archivos(archivo_A, archivo_B)

    # Leer el archivo existente para conservar las fórmulas para Datos_B
    wb = load_workbook(archivo_B)
    ws_B = wb["TAT"]

    # Obtener las fórmulas en la columna '% Pago' para Datos_B
    formulas_B = {cell.coordinate: cell for cell in ws_B["Y"][1:] if cell.data_type == "f"}
    formulas_B.update({cell.coordinate: cell for cell in ws_B["Z"][1:] if cell.data_type == "f"})
    formulas_B.update({cell.coordinate: cell for cell in ws_B["AG"][1:] if cell.data_type == "f"})
    formulas_B.update({cell.coordinate: cell for cell in ws_B["AH"][1:] if cell.data_type == "f"})
    formulas_B.update({cell.coordinate: cell for cell in ws_B["AI"][1:] if cell.data_type == "f"})
    formulas_B.update({cell.coordinate: cell for cell in ws_B["AJ"][1:] if cell.data_type == "f"})

    # Convertir los datos modificados de Datos_B a un dataframe de Pandas
    df_B = pd.DataFrame(datos_B)

    # Iterar sobre el DataFrame de Datos_B y escribir en el archivo respetando las fórmulas
    for r_idx, row in enumerate(dataframe_to_rows(df_B, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_B.cell(row=r_idx, column=c_idx)
            if cell.coordinate in formulas_B:
                ws_B[cell.coordinate] = formulas_B[cell.coordinate].value
            else:
                cell.value = value

    # Leer el archivo existente para conservar las fórmulas para Datos_Z
    ws_Z = wb["ICH"]

    # Obtener las fórmulas en la columna '% Pago' para Datos_Z
    formulas_Z = {cell.coordinate: cell for cell in ws_Z["Y"][1:] if cell.data_type == "f"}
    formulas_Z.update({cell.coordinate: cell for cell in ws_Z["Z"][1:] if cell.data_type == "f"})
    formulas_Z.update({cell.coordinate: cell for cell in ws_Z["AG"][1:] if cell.data_type == "f"})
    formulas_Z.update({cell.coordinate: cell for cell in ws_Z["AH"][1:] if cell.data_type == "f"})
    formulas_Z.update({cell.coordinate: cell for cell in ws_Z["AO"][1:] if cell.data_type == "f"})
    formulas_Z.update({cell.coordinate: cell for cell in ws_Z["AP"][1:] if cell.data_type == "f"})
    formulas_Z.update({cell.coordinate: cell for cell in ws_Z["AQ"][1:] if cell.data_type == "f"})
    formulas_Z.update({cell.coordinate: cell for cell in ws_Z["AR"][1:] if cell.data_type == "f"})
    
    # Convertir los datos modificados de Datos_Z a un dataframe de Pandas
    df_Z = pd.DataFrame(datos_Z)

    # Iterar sobre el DataFrame de Datos_Z y escribir en el archivo respetando las fórmulas
    for r_idx, row in enumerate(dataframe_to_rows(df_Z, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_Z.cell(row=r_idx, column=c_idx)
            if cell.coordinate in formulas_Z:
                ws_Z[cell.coordinate] = formulas_Z[cell.coordinate].value
            else:
                cell.value = value

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="LIQUIDACION KONQUISTADORES DICIEMBRE_modificado.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.register_blueprint(novedad_dic2023, url_prefix="/novedad")
    app.run(debug=True)