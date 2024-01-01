from flask import Blueprint, render_template, request, send_file
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

novedad_bp = Blueprint('novedad_bp', __name__)

def comparar_archivos(archivo_A, archivo_B):
    datos_A = pd.read_excel(archivo_A)
    datos_B = pd.read_excel(archivo_B, sheet_name='Datos_B')

    # Guardar las edades actuales en las columnas 'V - Edad' y 'V - Pago'
    datos_B['V - Edad'] = datos_B['Edad']
    datos_B['V - Pago'] = datos_B['Pago']

    for idx, fila_A in datos_A.iterrows():
        id_persona = fila_A['ID']
        edad_A = fila_A['Edad']
        pago_A = fila_A['Pago']

        if id_persona in datos_B['ID'].values:
            edad_B = datos_B.loc[datos_B['ID'] == id_persona, 'Edad'].values[0]
            pago_B = datos_B.loc[datos_B['ID'] == id_persona, 'Pago'].values[0]

            if edad_A != edad_B:
                datos_B.loc[datos_B['ID'] == id_persona, 'Edad'] = edad_A
            if pago_A != pago_B:
                datos_B.loc[datos_B['ID'] == id_persona, 'Pago'] = pago_A
        else:
            datos_B = pd.concat([datos_B, fila_A.to_frame().T], ignore_index=True)

    datos_B['R - Edad'] = 'Igual'
    datos_B['R - Pago'] = 'Igual'

    datos_B.loc[datos_B['V - Edad'] != datos_B['Edad'], 'R - Edad'] = 'Hubo cambio'
    datos_B.loc[datos_B['V - Pago'] != datos_B['Pago'], 'R - Pago'] = 'Hubo cambio'

    return datos_B

@novedad_bp.route('/')
def novedades():
    return render_template('/novedades/novedades.html')

@novedad_bp.route('/procesar_archivos', methods=['POST'])
def procesar_archivos():
    archivo_A = request.files['archivo_A']
    archivo_B = request.files['archivo_B']
    
    datos_modificados = comparar_archivos(archivo_A, archivo_B)

    # Leer el archivo existente para conservar las fórmulas
    wb = load_workbook(archivo_B)
    ws = wb['Datos_B']

    # Obtener las fórmulas en la columna '% Pago'
    formulas = {cell.coordinate: cell for cell in ws['K'][1:] if cell.data_type == 'f'}

    # Convertir los datos modificados a un dataframe de Pandas
    df = pd.DataFrame(datos_modificados)

    # Iterar sobre el DataFrame y escribir en el archivo respetando las fórmulas
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if cell.coordinate in formulas:
                ws[cell.coordinate] = formulas[cell.coordinate].value
            else:
                cell.value = value

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, as_attachment=True, download_name='Libro2_modificado.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(debug=True)